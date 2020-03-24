import xlsxwriter
import openpyxl


data9 = xlsxwriter.Workbook('data9.xlsx')
worksheet = data9.add_worksheet('data9.xlsx')
##worksheet = data9.add_worksheet('data10.xlsx')
worksheet.write('A1', 'Data 9')
row = 1
column = 0

content = ['Ben', 'GeeGee', 'Jonny', 'Cj']
for name in content:
    worksheet.write(row, column, name)
    row += 1

data9.close()
print(row)
print(column)
print(content)
print(worksheet)
# fetching excel


# path = 'data9.xlsx'
#
# data9_object = openpyxl.load_workbook(path)
#
# data9_object = data9_object.active
#
# get_data = data9_object.cell(row=1, column=1)
#
# total_row = data9_object.max_row
# total_column = data9_object.max_column
#
# # print(get_data)
# # print('Total columns = ', total_column)
# # print('Total rows = ', total_row)
#
# get_data.value = "steve"
# data9_object.cell(column=1, row=2).value = "Rodger"
#
# print(data9_object.cell(column=1, row=1).value)
# print(data9_object.cell(column=1, row=2).value)
#
# worksheet.save('data9.xlsx')
