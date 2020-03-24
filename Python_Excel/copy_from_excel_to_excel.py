import openpyxl as xl

filename = 'data9.xlsx'
wb1 = xl.load_workbook(filename)
ws1 = wb1.active

filename1 = 'copybook.xlsx'
wb2 = xl.load_workbook(filename)
ws2 = wb2.active

mr = ws1.max_row
mc = ws1.max_column

for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=i, column=j).value = c.value

wb2.save(str(filename1))

